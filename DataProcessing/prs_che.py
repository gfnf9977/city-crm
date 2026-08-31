import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from openpyxl.styles import PatternFill
from thefuzz import fuzz

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
}

types_to_check = [
    ("провулок", "пров."), ("пров.", "пров."),
    ("проспект", "просп."), ("просп.", "просп."),
    ("вулиця", "вул."), ("вул.", "вул."),
    ("узвіз", "узв."), ("узв.", "узв."),
    ("бульвар", "бульв."), ("бульв.", "бульв."),
    ("шосе", "шосе"), ("тупик", "тупик")
]


def parse_street(raw_text):
    raw_text = re.sub(r"[’‘`´]", "'", raw_text)
    raw_text = raw_text.replace('\xa0', ' ')

    clean = re.sub(r'\(.*?\)', '', raw_text)
    clean = re.sub(r'\[\d+\]', '', clean)
    clean = re.sub(r'[🟡🔵\u200b\n]', '', clean)

    lower_text = clean.lower()
    street_type = "вул."

    for full_type, short_type in types_to_check:
        if full_type in lower_text:
            street_type = short_type
            clean = re.sub(f'(?i){re.escape(full_type)}', '', clean)
            break

    clean_name = clean.strip(' ,-—"«».')
    clean_name = re.sub(r'\s+', ' ', clean_name).strip()
    return street_type, clean_name


def find_best_match(target_type, target_name, streets_list, threshold=90):
    best_match = None
    highest_score = 0

    for s in streets_list:
        if s["StreetType"] != target_type:
            continue

        name_db = s["Name"].lower()
        name_tgt = target_name.lower()

        score_ratio = fuzz.ratio(name_db, name_tgt)
        score_set = fuzz.token_set_ratio(name_db, name_tgt)

        score = max(score_ratio, score_set)

        if score > highest_score:
            highest_score = score
            best_match = s

    if highest_score >= threshold:
        return best_match, highest_score
    return None, 0


result_streets = []
added_names = set()
result_renames = []

print("1. Збираємо дані з Вікіпедії...")
wiki_url = "https://uk.wikipedia.org/wiki/Вулиці_Чернігова"
wiki_resp = requests.get(wiki_url, headers=headers)
wiki_soup = BeautifulSoup(wiki_resp.text, "html.parser")
content = wiki_soup.find("div", {"class": "mw-parser-output"})

for li in content.find_all("li"):
    text = li.get_text().strip()
    if "площа " in text.lower() or "сквер " in text.lower(): continue

    lower_text = text.lower()
    if not any(t[0] in lower_text for t in types_to_check): continue

    if re.search(r'\s+[-–—]\s+', text):
        parts = re.split(r'\s+[-–—]\s+', text)
        if len(parts) >= 2:
            old_raw, new_raw = parts[0], parts[1]
            if any(t[0] in old_raw.lower() or t[0] in new_raw.lower() for t in types_to_check):
                old_type, old_name = parse_street(old_raw)
                new_type, new_name = parse_street(new_raw)
                result_renames.append({
                    "Джерело": "Вікіпедія",
                    "Колишня назва": f"{old_type} {old_name}".strip(),
                    "OldType": old_type, "OldName": old_name,
                    "NewType": new_type, "NewName": new_name
                })
        continue

    if len(text) > 80: continue

    street_type, street_name = parse_street(text)
    unique_key = f"{street_type}_{street_name}"

    if street_type and street_name and unique_key not in added_names:
        result_streets.append({
            "StreetType": street_type, "Name": street_name,
            "OldNames": "", "_HighlightColor": None
        })
        added_names.add(unique_key)

print("2. Збираємо нові перейменування з Міськради...")
rada_url = "https://chernigiv-rada.gov.ua/vulytsi-ta-provulky/"
rada_resp = requests.get(rada_url, headers=headers)
rada_soup = BeautifulSoup(rada_resp.text, "html.parser")

for table in rada_soup.find_all("table"):
    for row in table.find_all("tr")[1:]:
        cols = row.find_all("td")
        if len(cols) >= 2:
            old_raw, new_raw = cols[0].get_text().strip(), cols[1].get_text().strip()
            old_type, old_name = parse_street(old_raw)
            new_type, new_name = parse_street(new_raw)

            is_duplicate = False
            for r in result_renames:
                if r["NewType"] == new_type and fuzz.ratio(r["NewName"].lower(), new_name.lower()) > 90:
                    is_duplicate = True
                    break

            if old_name and new_name and not is_duplicate:
                result_renames.append({
                    "Джерело": "Міськрада",
                    "Колишня назва": f"{old_type} {old_name}".strip(),
                    "OldType": old_type, "OldName": old_name,
                    "NewType": new_type, "NewName": new_name
                })

print("\n3. Розумне злиття (Fuzzy Matching)...")
print("-" * 50)


def add_old_name(street_dict, new_old_name):
    if street_dict["OldNames"]:
        if new_old_name not in street_dict["OldNames"]:
            street_dict["OldNames"] = f"{new_old_name}, {street_dict['OldNames']}"
    else:
        street_dict["OldNames"] = new_old_name


for rename in result_renames:
    matched_new, score_new = find_best_match(rename["NewType"], rename["NewName"], result_streets, threshold=90)

    if matched_new:
        if score_new < 100:
            print(f"🤖 Знайдено збіг: '{rename['NewName']}' ~ '{matched_new['Name']}' (Точність: {score_new}%)")
        add_old_name(matched_new, rename["Колишня назва"])

    else:
        matched_old, score_old = find_best_match(rename["OldType"], rename["OldName"], result_streets, threshold=90)

        if matched_old:
            if score_old < 100:
                print(f"🤖 Автовиправлення: '{rename['OldName']}' ~ '{matched_old['Name']}' (Точність: {score_old}%)")

            matched_old["_HighlightColor"] = "Yellow"
            add_old_name(matched_old, rename["Колишня назва"])
            matched_old["StreetType"] = rename["NewType"]
            matched_old["Name"] = rename["NewName"]

        else:
            new_street = {
                "StreetType": rename["NewType"],
                "Name": rename["NewName"],
                "OldNames": rename["Колишня назва"],
                "_HighlightColor": "Green"
            }
            result_streets.append(new_street)

print("-" * 50)

print("4. Зберігаємо в Excel з підсвіткою клітинок...")
colors_list = [s.get('_HighlightColor') for s in result_streets]
for s in result_streets:
    s.pop('_HighlightColor', None)

for r in result_renames:
    r.pop('OldType', None);
    r.pop('OldName', None)
    r.pop('NewType', None);
    r.pop('NewName', None)

df_streets = pd.DataFrame(result_streets)

with pd.ExcelWriter('streets_data.xlsx', engine='openpyxl') as writer:
    df_streets.to_excel(writer, sheet_name='Вулиці', index=False)
    worksheet = writer.sheets['Вулиці']
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    for idx, color in enumerate(colors_list):
        if color:
            fill = yellow_fill if color == "Yellow" else green_fill
            for col in range(1, len(df_streets.columns) + 1):
                worksheet.cell(row=idx + 2, column=col).fill = fill

    if result_renames:
        pd.DataFrame(result_renames).to_excel(writer, sheet_name='Перейменування', index=False)

print("Готово!")